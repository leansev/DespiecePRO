# -*- coding: utf-8 -*-
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    'cantidad',
    'LARGO',
    'ANCHO',
    'nombre',
    'rota',
    'canto_arr',
    'canto_aba',
    'canto_izq',
    'canto_der',
]

COLOR_HEADER_BG = '2F4F7F'
COLOR_HEADER_FG = 'FFFFFF'
COLOR_MODULE_BG = '4A7C9E'
COLOR_MODULE_FG = 'FFFFFF'
COLOR_ROW_ALT = 'EEF2F7'
COLOR_WHITE = 'FFFFFF'


def solid_fill(color):
    return PatternFill(fill_type='solid', fgColor=color)


def load_payload(json_path):
    with open(json_path, encoding='utf-8-sig') as handle:
        return json.load(handle)


def group_modules(rows):
    modules = []
    current_module = None
    current_pieces = []

    for item in rows:
        row_type = item.get('type')
        if row_type == 'module':
            if current_module is not None:
                modules.append((current_module, current_pieces))
            current_module = item.get('label', '')
            current_pieces = []
        elif row_type == 'piece':
            current_pieces.append(item)

    if current_module is not None:
        modules.append((current_module, current_pieces))

    return modules


def collect_thicknesses(modules):
    thicknesses = set()
    for _, pieces in modules:
        for piece in pieces:
            thicknesses.add(piece.get('espesor', 0))
    return sorted(thicknesses, reverse=True)


def sheet_title(thickness):
    if thickness == int(thickness):
        return f'{int(thickness)}mm'
    return f'{thickness}mm'


def write_header_row(sheet, row_index):
    fill = solid_fill(COLOR_HEADER_BG)
    font = Font(bold=True, color=COLOR_HEADER_FG)

    for column_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')


def write_module_row(sheet, row_index, label):
    fill = solid_fill(COLOR_MODULE_BG)
    font = Font(bold=True, color=COLOR_MODULE_FG)

    for column_index in range(1, len(HEADERS) + 1):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.fill = fill
        if column_index == 1:
            cell.value = label
            cell.font = font

    sheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=len(HEADERS),
    )


def write_piece_row(sheet, row_index, item, use_alt_fill):
    values = [
        item['cantidad'],
        item['largo'],
        item['ancho'],
        item['nombre'],
        item.get('rota', 1),
        item.get('canto_arr', 0),
        item.get('canto_aba', 0),
        item.get('canto_izq', 0),
        item.get('canto_der', 0),
    ]
    fill_color = COLOR_ROW_ALT if use_alt_fill else COLOR_WHITE
    fill = solid_fill(fill_color)

    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        cell.fill = fill


def write_total_row(sheet, row_index, label):
    cell = sheet.cell(row=row_index, column=1, value=label)
    cell.font = Font(bold=True)


def adjust_column_widths(sheet):
    for column_index, header in enumerate(HEADERS, start=1):
        width = max(len(header) + 2, 12)
        column_letter = sheet.cell(row=2, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = width


def write_sheet(workbook, sheet_name, project_title, modules, thickness):
    sheet = workbook.create_sheet(title=sheet_name)

    row_index = 1
    title_cell = sheet.cell(row=row_index, column=1, value=project_title)
    title_cell.font = Font(bold=True)
    sheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=len(HEADERS),
    )
    row_index += 1

    write_header_row(sheet, row_index)
    row_index += 1

    total_count = 0
    for module_label, pieces in modules:
        filtered_pieces = [
            piece for piece in pieces if piece.get('espesor') == thickness
        ]
        if not filtered_pieces:
            continue

        write_module_row(sheet, row_index, module_label)
        row_index += 1

        use_alt_fill = False
        for piece in filtered_pieces:
            write_piece_row(sheet, row_index, piece, use_alt_fill)
            total_count += piece.get('cantidad', 0)
            use_alt_fill = not use_alt_fill
            row_index += 1

    write_total_row(sheet, row_index, f'TOTAL DE PIEZAS: {total_count}')
    adjust_column_widths(sheet)


def write_xlsx(output_path, payload):
    project_title = payload.get('project_title', 'PROYECTO: Sin nombre')
    modules = group_modules(payload.get('rows', []))
    thicknesses = collect_thicknesses(modules)

    if not thicknesses:
        thicknesses = [0]

    workbook = Workbook()
    workbook.remove(workbook.active)

    for thickness in thicknesses:
        write_sheet(
            workbook,
            sheet_title(thickness),
            project_title,
            modules,
            thickness,
        )

    workbook.save(output_path)


def main():
    if len(sys.argv) < 3:
        sys.stderr.write('Uso: export_excel.py salida.xlsx datos.json\n')
        sys.exit(1)

    output_path = sys.argv[1]
    json_path = sys.argv[2]
    payload = load_payload(json_path)
    write_xlsx(output_path, payload)


if __name__ == '__main__':
    main()
